# ChatFlow server




import socket
import threading
import openpyxl
import datetime



HOST = ''
PORT = 1234


userdataworkbook = openpyxl.load_workbook('userdata.xlsx')
userdata = userdataworkbook.active

chathistoryworkbook = openpyxl.load_workbook('chathistory.xlsx')
chathistory = chathistoryworkbook.active

server = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
server.bind((HOST, PORT))
server.listen()


clients = dict() # client socket mapped to username




def broadcast(message):

    """Send message to all connected clients."""
    dtnow = datetime.datetime.now()
    time = dtnow.strftime('%H:%M')
    message = message.decode()
    message = '\n' + '{} | '.format(time) + message
    message = message.encode()

    for client in clients.keys():
        client.send(message)




def handle(client):
    """Manage messages from a single client."""

    while True:
        try:
            message = client.recv(1024)
            if message.decode() == 'File Transfer':
                client.send('Send File Name'.encode())
                fileName = client.recv(1024).decode()
                print(fileName)

                client.send('Send File'.encode())
                file = open(fileName,'wb')
                l = client.recv(1024)

                while(l):
                    print('inside loop')
                    file.write(l)
                    l = client.recv(1024)
                    if l  == 'Completed'.encode():
                        break
                    
                print('outside loop')
                file.close()


            else:
                message = message.decode()
                user = message.split(":")[0].strip()
                mssg = message.split(":")[1].strip()
                i = chathistory.max_row + 1

                chathistory.cell(i,1).value = datetime.datetime.now()
                chathistory.cell(i,2).value = user
                chathistory.cell(i,3).value = mssg

                message = message.encode()
                broadcast(message)

        except:
            username = clients.pop(client)
            broadcast('{} left the chat!'.format(username).encode())
            break




def login(client, username, password):
    """Authenticate an existing user."""

    for i in range(1, userdata.max_row + 1):
        if userdata.cell(i, 1).value == username:
            if userdata.cell(i, 3).value == password:
                return True

            else:
                client.send('Incorrect Password'.encode())
                return False

    client.send('User does not exist'.encode())
    return False






def register(client, username, email, password):
    """Create a new account."""

    i = userdata.max_row + 1


    for i in range(1, userdata.max_row + 1):

        if userdata.cell(i,1).value == username:
            client.send('User already exist'.encode())
            return False

    userdata.cell(i, 1).value = username
    userdata.cell(i, 2).value = email
    userdata.cell(i, 3).value = password

    userdataworkbook.save('userdata.xlsx')
    return True




while True:

    client, address = server.accept()
    print('Connected with {}'.format(address))


    client.send('Login or Reg'.encode())
    auth_mode = client.recv(1024).decode()




    if auth_mode == 'Login':
        client.send('USER'.encode())
        username = client.recv(1024).decode()

        client.send('PW'.encode())
        password = client.recv(1024).decode()

        authenticated = login(client,username, password)




    elif auth_mode == 'Register':
        client.send('USER'.encode())
        username = client.recv(1024).decode()

        client.send('PW'.encode())
        password = client.recv(1024).decode()

        client.send('EMAIL'.encode())
        email = client.recv(1024).decode()

        authenticated = register(client, username, email, password)




    if authenticated == True:

        client.send('Authenticated'.encode())
        clients[client] = username
        print('Client Username: {}'.format(username))

        for i in range(2, chathistory.max_row + 1):

            time = chathistory.cell(i,1).value.strftime('%H:%M')
            user = chathistory.cell(i,2).value
            mssg = chathistory.cell(i,3).value

            client.send('\n{} | {} : {}'.format(time,user,mssg).encode())

        broadcast('{} joined the Chat!'.format(username).encode())
        threading.Thread(target = handle, args = (client,)).start()


    else:
        client.send('Authentication Failed'.encode())



